"""
excel_api.py

Minimal abstraction for reading an Excel workbook and exposing its sheets.

PUBLIC API (import these only):

    - ExcelError
    - ExcelSheetNotFound
    - ExcelSheetInfo
    - ExcelWorkbook
    - load_excel_workbook(path)

This module intentionally hides the concrete Excel backend (openpyxl)
so that the rest of the application depends ONLY on this file.
"""

from __future__ import annotations


from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Any, Iterable, Optional, Tuple

__all__ = [
    "ExcelError",
    "ExcelSheetNotFound",
    "ExcelSheetInfo",
    "ExcelWorkbook",
    "load_excel_workbook",
    "save_excel_workbook"
]


# =========================
# Exceptions (Public)
# =========================

class ExcelError(Exception):
    """Base exception for all Excel-related errors in this module."""
    pass


class ExcelSheetNotFound(ExcelError):
    """Raised when a requested sheet does not exist in the workbook."""

    def __init__(self, sheet_name: str, available: Iterable[str]) -> None:
        available_list = list(available)
        msg = (
            f"Sheet '{sheet_name}' not found in workbook. "
            f"Available sheets: {available_list}"
        )
        super().__init__(msg)
        self.sheet_name = sheet_name
        self.available = available_list


# =========================
# Data model (Public)
# =========================
@dataclass(frozen=True)
class ExcelSheetInfo:
    """
    Basic metadata about a single sheet.

    Attributes:
        name: Sheet name
        n_rows: Number of rows (including empty rows if present in the 2D data)
        n_cols: Maximum number of columns in any row
    """
    name: str
    n_rows: int
    n_cols: int



@dataclass(frozen=True)
class ExcelWorkbook:
    """
    Immutable representation of an Excel workbook.

    Internally stores:
        _sheets: Dict[sheet_name, List[List[Any]]]
        _sheet_info: Dict[sheet_name, ExcelSheetInfo]
    """

    path: Path
    _sheets: Dict[str, List[List[Any]]]
    _sheet_info: Dict[str, ExcelSheetInfo]

    # -------- Public methods --------

    def sheet_names(self) -> List[str]:
        """Return list of sheet names in the workbook."""
        return list(self._sheets.keys())

    def sheet_count(self) -> int:
        """Return the number of sheets in the workbook."""
        return len(self._sheets)

    def has_sheet(self, name: str) -> bool:
        """Check if a sheet with this name exists."""
        return name in self._sheets

    def get_sheet_values(self, name: str) -> List[List[Any]]:
        """
        Return the sheet content as a 2D list (rows of cells).

        Raises:
            ExcelSheetNotFound
        """
        try:
            return self._sheets[name]
        except KeyError:
            raise ExcelSheetNotFound(name, self._sheets.keys())

    def iter_sheet_values(self, name: str) -> Iterable[List[Any]]:
        """Iterate over the rows of the given sheet."""
        return iter(self.get_sheet_values(name))

    def first_sheet_name(self) -> Optional[str]:
        """Return the first sheet name, or None if workbook empty."""
        names = self.sheet_names()
        return names[0] if names else None

    # -------- New: dimension-related helpers --------

    def get_sheet_info(self, name: str) -> ExcelSheetInfo:
        """
        Return ExcelSheetInfo for the given sheet.

        Raises:
            ExcelSheetNotFound
        """
        try:
            return self._sheet_info[name]
        except KeyError:
            raise ExcelSheetNotFound(name, self._sheets.keys())

    def sheet_dimensions(self, name: str) -> Tuple[int, int]:
        """
        Convenience: return (n_rows, n_cols) for the given sheet.

        Raises:
            ExcelSheetNotFound
        """
        info = self.get_sheet_info(name)
        return info.n_rows, info.n_cols



# =========================
# Loader function (Public)
# =========================

def load_excel_workbook(path: str | Path) -> ExcelWorkbook:
    """
    Load an Excel workbook from disk and return an ExcelWorkbook instance.
    """
    p = Path(path)

    if not p.is_file():
        raise ExcelError(f"Excel file does not exist: {p}")

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ExcelError(
            "openpyxl is required to read Excel files. "
            "Please install it with: pip install openpyxl"
        ) from exc

    try:
        wb = load_workbook(p, data_only=True, read_only=True)
    except Exception as exc:
        raise ExcelError(f"Failed to open Excel file '{p}': {exc}") from exc

    sheets: Dict[str, List[List[Any]]] = {}
    sheet_info: Dict[str, ExcelSheetInfo] = {}

    try:
        for ws in wb.worksheets:
            rows_2d: List[List[Any]] = []
            max_cols = 0

            for row in ws.iter_rows(values_only=True):
                row_list = list(row)
                rows_2d.append(row_list)
                if len(row_list) > max_cols:
                    max_cols = len(row_list)

            name = ws.title
            sheets[name] = rows_2d

            info = ExcelSheetInfo(
                name=name,
                n_rows=len(rows_2d),
                n_cols=max_cols,
            )
            sheet_info[name] = info

    except Exception as exc:
        raise ExcelError(f"Failed to read sheets from Excel file '{p}': {exc}") from exc
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return ExcelWorkbook(path=p, _sheets=sheets, _sheet_info=sheet_info)






def save_excel_workbook(workbook: ExcelWorkbook, path: str | Path) -> None:
    """
    Save the given ExcelWorkbook to disk as a real .xlsx file.

    - Writes ALL sheets in the workbook.
    - Each sheet keeps its original name.
    - Cell values are written as-is from the 2D lists.

    Args:
        workbook: ExcelWorkbook instance to export.
        path: Target file path (.xlsx).

    Raises:
        ExcelError: if openpyxl is missing or writing fails.
    """
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError as exc:
        raise ExcelError(
            "openpyxl is required to write Excel files. "
            "Please install it with: pip install openpyxl"
        ) from exc

    p = Path(path)

    try:
        sheet_names = workbook.sheet_names()

        wb = Workbook()
        active_ws = wb.active

        if sheet_names:
            # Use the active sheet for the first sheet in our workbook
            first_name = sheet_names[0]
            active_ws.title = first_name

            # Fill first sheet
            rows = workbook.get_sheet_values(first_name)
            for r_idx, row in enumerate(rows, start=1):
                if row is None:
                    continue
                for c_idx, value in enumerate(row, start=1):
                    active_ws.cell(row=r_idx, column=c_idx, value=value)

            # Create remaining sheets
            for name in sheet_names[1:]:
                ws = wb.create_sheet(title=name)
                rows = workbook.get_sheet_values(name)
                for r_idx, row in enumerate(rows, start=1):
                    if row is None:
                        continue
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
        else:
            # No sheets in ExcelWorkbook:
            # leave default empty sheet with default name.
            pass

        wb.save(p)

    except Exception as exc:
        raise ExcelError(f"Failed to write Excel file '{p}': {exc}") from exc
